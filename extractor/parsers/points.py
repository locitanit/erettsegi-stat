"""Pontszamok a pontozotablabol (`*_ertekelo_*.xlsx`).

Miert nem a PDF-bo"l? Az utmutato-PDF-ben a pontszamok keretezett dobozokban
allnak, es a szovegkinyeres elcsusztatja o"ket: a "12 pont" gyakran egy masik
sorba kerul. A pontozotabla xlsx ugyanezt strukturaltan tartalmazza:

    B oszlop = a pontozasi sor szovege
    C oszlop = az adhato maximalis pont

Ket szint van benne, es MINDKETTO kell:

1. A tabla **vegen egy osszesito**, feladatonkent egy sorral es a vegen a
   vizsga osszpontszamaval - ez a vizsgapont megbizhato forrasa:

       5. Létra      15
                    100

2. A tabla **torzse**, feladatonkent a reszfeladatok pontjaival es a
   "Feladatpontok összesen" sorral - ez a feladatpont es a reszfeladatszam.

2012 elo"tt nincs ilyen xlsx (csak PDF-ertekelo"lap), ott a pontszam ismeretlen marad.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

# "1. Az Árpád motorkocsi", "1A. Síparadicsomok", "1.A Dobogókő"
TASK_HEADER = re.compile(
    r"^\s*(?P<num>\d)(?:\.\s*(?P<sub1>[AB])\b|(?P<sub2>[AB])\.|\.)\s*(?P<title>\S.{0,60})$"
)
TOTAL_ROW = re.compile(r"(feladatpontok?\s*összesen|^\s*összesen)\s*:?", re.IGNORECASE)
EXAM_POINT_ROW = re.compile(r"vizsgapont", re.IGNORECASE)


@dataclass
class TaskPoints:
    task_no: str
    title: str
    points: int | None = None          # feladatpont (a pontozas sajat skalaja)
    exam_points: int | None = None     # vizsgapont (amit a vizsgan er)
    subtasks: list[int] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "points": self.points,
            "exam_points": self.exam_points,
            "subtask_count": len(self.subtasks),
            "subtask_points": self.subtasks,
        }


def _number(value: object) -> int | None:
    """Egesz pontszam a cellabol. A "15 pont" alaku szoveget is elfogadja."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and float(value).is_integer() and 0 < value <= 300:
        return int(value)
    if isinstance(value, str):
        m = re.fullmatch(r"\s*(\d{1,3})\s*pont\s*", value, re.IGNORECASE)
        if m:
            return int(m.group(1))
    return None


def _rows(sheet) -> list[tuple[str, int | None]]:
    """(szoveg, pontszam) parok. A szoveg az elso nem ures szoveges cella."""
    out: list[tuple[str, int | None]] = []
    for row in sheet.iter_rows(values_only=True):
        cells = list(row)
        text = next((str(c) for c in cells if isinstance(c, str) and c.strip()), "")
        number = next((n for n in (_number(c) for c in cells) if n is not None), None)
        out.append((" ".join(text.split()), number))
    return out


def _summary_block(rows: list[tuple[str, int | None]]) -> tuple[int, dict[str, int]]:
    """A tabla vegi osszesito: (kezdo sorindex, feladatszam -> vizsgapont).

    Legalabb ket egymast koveto feladat-sor kell hozza, mindegyiken pontszammal.
    Ha nincs ilyen, (len(rows), {}) jon vissza.
    """
    i = len(rows) - 1
    while i >= 0:
        text, number = rows[i]
        m = TASK_HEADER.match(text) if text else None
        if not (m and number is not None):
            i -= 1
            continue
        # visszafele osszegyujtjuk az egybefuggo feladat-sorokat
        block: dict[str, int] = {}
        j = i
        while j >= 0:
            t, n = rows[j]
            mm = TASK_HEADER.match(t) if t else None
            if mm and n is not None and not TOTAL_ROW.search(t) and not EXAM_POINT_ROW.search(t):
                block[f"{mm['num']}{mm['sub1'] or mm['sub2'] or ''}"] = n
                j -= 1
            elif not t:
                break
            else:
                break
        if len(block) >= 2:
            return j + 1, block
        i = j - 1
    return len(rows), {}


def read_scoring_sheet(path: Path) -> tuple[dict[str, TaskPoints], list[str]]:
    """Egy pontozotabla feldolgozasa: feladatszam -> pontok."""
    warnings: list[str] = []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True, read_only=True)
    except ImportError:
        return {}, ["az openpyxl nincs telepítve, a pontszámok kimaradtak"]
    except Exception as exc:
        return {}, [f"{path.name}: nem olvasható ({type(exc).__name__})"]

    tasks: dict[str, TaskPoints] = {}
    try:
        # A tanari lap ("Vizsgazo1") a leghosszabb; a hasznalati utmutato rovid.
        sheet = max(wb.worksheets, key=lambda ws: ws.max_row or 0)
        rows = _rows(sheet)
        summary_start, summary = _summary_block(rows)

        current: TaskPoints | None = None
        for text, number in rows[:summary_start]:
            if not text:
                continue
            header = TASK_HEADER.match(text)
            # Feladat-fejlec csak az a sor lehet, amelyhez NINCS pontszam: a
            # pontozasi sorok kulonben fejlecnek latszananak.
            if (
                header
                and number is None
                and not TOTAL_ROW.search(text)
                and not EXAM_POINT_ROW.search(text)
            ):
                key = f"{header['num']}{header['sub1'] or header['sub2'] or ''}"
                current = tasks.setdefault(
                    key, TaskPoints(task_no=key, title=header["title"].strip())
                )
                continue
            if current is None or number is None:
                continue
            if EXAM_POINT_ROW.search(text):
                current.exam_points = number
            elif TOTAL_ROW.search(text):
                current.points = number
            else:
                current.subtasks.append(number)
    except Exception as exc:
        warnings.append(f"{path.name}: hiba olvasás közben ({type(exc).__name__})")
        summary = {}
    finally:
        wb.close()

    for key, value in summary.items():
        task = tasks.setdefault(key, TaskPoints(task_no=key, title=""))
        task.exam_points = value          # az osszesito az elsodleges forras
    for t in tasks.values():
        if t.points is None and t.subtasks:
            t.points = sum(t.subtasks)
        if t.exam_points is None:
            t.exam_points = t.points
    return tasks, warnings


def exam_total(tasks: dict[str, TaskPoints]) -> int:
    """A vizsga osszpontszama. Az azonos sorszamu A/B feladat valaszthato: egyszer szamit."""
    by_number: dict[str, int] = {}
    for key, task in tasks.items():
        number = key.rstrip("AB")
        by_number[number] = max(by_number.get(number, 0), task.exam_points or 0)
    return sum(by_number.values())


def read_exam_points(paths: list[Path]) -> tuple[dict[str, TaskPoints], list[str]]:
    """A vizsga pontozotablai kozul a legjobb: eloszor a helyes osszpontszamu."""
    candidates: list[dict[str, TaskPoints]] = []
    warnings: list[str] = []
    for path in paths:
        tasks, w = read_scoring_sheet(path)
        warnings += w
        if tasks:
            candidates.append(tasks)
    if not candidates:
        return {}, warnings
    valid = [c for c in candidates if exam_total(c) in (100, 120)]
    return (valid or candidates)[0], warnings
