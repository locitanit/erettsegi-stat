"""Tablazatkezeles-elemzo: fuggvenyek, egyuttes elofordulas, keplet-komplexitas, keszsegek.

Ket forrasbol dolgozik (TERV 1.2/4):
  - **utmutato-PDF szovege** (elsodleges): itt vannak a mintakepletek
  - **mintamegoldas-xlsx** (megerosites): openpyxl-lel a kepletek angol nevekkel

A ketto kulon mezoben marad, hogy a feluleten latszodjon, honnan jott az adat.
"""
from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from itertools import combinations
from pathlib import Path

from ..vocab_store import load_functions, load_keywords
from .common import count_keywords, formula_candidates, max_paren_depth

# Minden "NÉV(" token, zarojel-melysegtol fuggetlenul (TERV 1.2/5).
# A nev nagybetuvel kezdodik, tartalmazhat pontot (KEREK.LE, AB.SZUM, HOL.VAN).
FUNC_TOKEN = re.compile(
    r"(?<![A-Za-zÁÉÍÓÖŐÚÜŰáéíóöőúüű0-9._])"
    r"([A-ZÁÉÍÓÖŐÚÜŰ][A-ZÁÉÍÓÖŐÚÜŰ0-9._]{1,24})\s*\("
)

# Ezek soha nem fuggvenynevek. Az IF/AND/OR/NOT szandekosan NINCS itt: azok
# valodi angol aliasok (HA/ES/VAGY/NEM), es az xlsx-ben mindig igy szerepelnek.
IGNORE_TOKENS = frozenset(
    {"RGB", "SELECT", "FROM", "WHERE", "INTO", "GROUP", "ORDER", "JOIN", "ON", "AS", "VALUES"}
)

# Cellahivatkozas, ami veletlenul zarojel ele kerult: "=B3 (ekkor a C22-es cella ures)".
CELL_REF = re.compile(r"^\$?[A-Z]{1,3}\$?\d{1,7}$")


@dataclass
class FormulaStats:
    """Egy forrasbol (utmutato vagy xlsx) kinyert keplet-statisztika."""
    functions: Counter = field(default_factory=Counter)     # kanonikus nev -> hany kepletben
    pairs: Counter = field(default_factory=Counter)         # "A+B" -> hany kepletben egyutt
    unknown: Counter = field(default_factory=Counter)
    formula_count: int = 0
    depth_max: int = 0
    functions_per_formula_max: int = 0

    def as_dict(self) -> dict:
        return {
            "functions": dict(self.functions.most_common()),
            "function_pairs": dict(self.pairs.most_common()),
            "formula_count": self.formula_count,
            "formula_depth_max": self.depth_max,
            "functions_per_formula_max": self.functions_per_formula_max,
        }


def _tokens(formula: str) -> tuple[list[str], list[str]]:
    """A kepletben szereplo fuggvenyek kanonikus nevei + az ismeretlen tokenek."""
    fns = load_functions()
    known: list[str] = []
    unknown: list[str] = []
    for m in FUNC_TOKEN.finditer(formula):
        token = m.group(1).rstrip(".")
        if token in IGNORE_TOKENS or len(token) < 2 or CELL_REF.match(token):
            continue
        canon = fns.canonical(token)
        if canon:
            known.append(canon)
        else:
            unknown.append(token)
    return known, unknown


def analyse_formulas(formulas: list[str]) -> FormulaStats:
    st = FormulaStats()
    for f in formulas:
        known, unknown = _tokens(f)
        for u in unknown:
            st.unknown[u] += 1
        if not known:
            continue
        st.formula_count += 1
        distinct = sorted(set(known))
        for name in distinct:
            st.functions[name] += 1
        for a, b in combinations(distinct, 2):
            st.pairs[f"{a}+{b}"] += 1
        st.depth_max = max(st.depth_max, max_paren_depth(f))
        st.functions_per_formula_max = max(st.functions_per_formula_max, len(known))
    return st


def from_utmutato(section_text: str) -> tuple[FormulaStats, dict[str, int]]:
    """Az utmutato-szakaszbol: keplet-statisztika + keszseg-kulcsszavak."""
    stats = analyse_formulas(formula_candidates(section_text))
    skills = count_keywords(section_text, load_keywords("tablazat_skills.yaml"))
    return stats, skills


def from_xlsx(paths: list[Path]) -> tuple[FormulaStats, list[str]]:
    """A mintamegoldas-xlsx kepletei. Hiba eseten ures statisztika + figyelmeztetes."""
    warnings: list[str] = []
    formulas: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return FormulaStats(), ["az openpyxl nincs telepítve, az xlsx-ek kimaradtak"]

    for path in paths:
        if path.suffix.lower() != ".xlsx":
            continue
        try:
            wb = load_workbook(str(path), data_only=False, read_only=True)
        except Exception as exc:                       # serult vagy jelszavas fajl
            warnings.append(f"{path.name}: nem olvasható ({type(exc).__name__})")
            continue
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.startswith("="):
                            formulas.append(v)
        except Exception as exc:
            warnings.append(f"{path.name}: hiba olvasás közben ({type(exc).__name__})")
        finally:
            wb.close()
    return analyse_formulas(formulas), warnings
